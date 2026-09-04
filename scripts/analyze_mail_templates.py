"""Read-only clustering of archived mail and attachment layouts by sender domain."""

import argparse
import io
import json
import math
import re
import zipfile
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from pathlib import Path

import xlrd
from openpyxl import load_workbook
from sqlalchemy import create_engine, text

from app.parsing import ALIAS_MAP


@dataclass
class Scan:
    domain: str
    filename: str
    extension: str
    sheets: list[str] = field(default_factory=list)
    fields: dict[str, set[str]] = field(default_factory=lambda: defaultdict(set))
    positions: set[str] = field(default_factory=set)
    labels: set[str] = field(default_factory=set)
    error: str | None = None


def clean(value):
    return re.sub(r"\s+", "", str(value or "")).strip()


def domain_of(value):
    match = re.search(r"@([A-Za-z0-9.-]+)", value or "")
    return match.group(1).lower() if match else "unknown"


def normalize_pattern(value):
    value = re.sub(r"20\d{2}[-/.年]?\d{1,2}[-/.月]?\d{1,2}日?", "{日期}", value)
    value = re.sub(r"(?<![A-Za-z0-9])[A-Z]{1,4}\d{3,8}(?:\([^)]+\))?", "{产品代码}", value)
    value = re.sub(r"\d{8,}", "{编号}", value)
    return value[:180]


def likely_label(value):
    value = clean(value)
    if not value or len(value) > 60:
        return False
    if re.fullmatch(r"[-+,.%\d/年月日: ]+", value):
        return False
    return bool(re.search(r"[\u4e00-\u9fffA-Za-z]", value))


def scan_rows(scan, sheet_index, sheet_name, rows):
    scan.sheets.append(sheet_name)
    for row_index, row in enumerate(rows[:80], 1):
        for column_index, value in enumerate(row[:80], 1):
            normalized = clean(value).lower()
            field_name = ALIAS_MAP.get(normalized)
            if field_name:
                label = clean(value)
                scan.fields[field_name].add(label)
                scan.positions.add(f"S{sheet_index}:R{row_index}C{column_index}:{field_name}")
            if likely_label(value):
                scan.labels.add(clean(value))


def scan_book(domain, filename, content):
    extension = Path(filename).suffix.lower()
    scan = Scan(domain=domain, filename=filename, extension=extension)
    try:
        if extension in {".xlsx", ".xlsm"}:
            if extension == ".xlsm":
                raise ValueError("含宏工作簿不自动分析")
            book = load_workbook(
                io.BytesIO(content), read_only=True, data_only=True, keep_links=False
            )
            try:
                for sheet_index, sheet in enumerate(book, 1):
                    rows = list(sheet.iter_rows(max_row=80, max_col=80, values_only=True))
                    scan_rows(scan, sheet_index, sheet.title, rows)
            finally:
                book.close()
        elif extension == ".xls":
            book = xlrd.open_workbook(file_contents=content, on_demand=True)
            try:
                for sheet_index, sheet in enumerate(book.sheets(), 1):
                    rows = [
                        [cell.value for cell in sheet.row(row_index)[:80]]
                        for row_index in range(min(sheet.nrows, 80))
                    ]
                    scan_rows(scan, sheet_index, sheet.name, rows)
            finally:
                book.release_resources()
        else:
            scan.error = "非 Excel 附件"
    except Exception as exc:
        scan.error = f"{type(exc).__name__}: {str(exc)[:160]}"
    return scan


def iter_attachment_files(filename, content):
    if Path(filename).suffix.lower() != ".zip":
        yield filename, content
        return
    try:
        with zipfile.ZipFile(io.BytesIO(content)) as archive:
            members = [m for m in archive.infolist() if not m.is_dir()]
            if len(members) > 100 or sum(m.file_size for m in members) > 100 * 1024 * 1024:
                return
            for member in members:
                safe_name = Path(member.filename.replace("\\", "/")).name
                if safe_name and Path(safe_name).suffix.lower() in {".xlsx", ".xls"}:
                    yield f"{filename} → {safe_name}", archive.read(member)
    except (OSError, zipfile.BadZipFile):
        return


def signature(scan):
    normalized_sheets = tuple(normalize_pattern(name) for name in scan.sheets)
    return scan.extension, normalized_sheets, tuple(sorted(scan.positions))


def render_report(originals, attachments, scans, statuses, job_results, output):
    lines = [
        "# 邮箱托管模板与可解析表头清单",
        "",
        "> 生成方式：只读扫描归档邮件与附件；按发件域和固定版式聚类，不逐封建立解析规则。",
        "",
        "## 一、扫描概况",
        "",
        f"- 原始邮件：{len(originals)} 封；附件：{len(attachments)} 个；可读取 Excel 文件（含 ZIP 内文件）：{sum(not s.error for s in scans)} 个。",
        f"- 解析任务状态：{'、'.join(f'{key} {value}' for key, value in sorted(statuses.items())) or '无'}。",
        "- 托管模板以发件域为第一层边界；同域内再按工作表名称、已识别字段位置和附件类型聚类。",
        "",
        "## 二、邮件头模式",
        "",
        "系统可稳定读取并归档：发件人、主题、Message-ID、接收时间、邮箱文件夹以及附件名称。下表按发件域汇总主题模板。",
        "",
        "| 发件域 | 邮件数 | 附件数 | 主要主题模式 |",
        "|---|---:|---:|---|",
    ]
    original_by_domain = Counter(domain_of(item["from"]) for item in originals)
    attachment_by_domain = Counter(item["domain"] for item in attachments)
    subjects = defaultdict(Counter)
    for item in originals:
        subjects[domain_of(item["from"])][normalize_pattern(item["subject"])] += 1
    domains = sorted(set(original_by_domain) | set(attachment_by_domain))
    for domain in domains:
        patterns = (
            "人工来信/非托管模板"
            if domain == "163.com"
            else "邮箱服务系统通知"
            if domain == "service.netease.com"
            else "<br>".join(
                f"{pattern}（{count}）"
                for pattern, count in subjects[domain].most_common(4)
            )
        )
        lines.append(
            f"| {domain} | {original_by_domain[domain]} | {attachment_by_domain[domain]} | {patterns or '—'} |"
        )

    lines += ["", "## 三、附件模板与可解析字段", ""]
    scans_by_domain = defaultdict(list)
    for scan in scans:
        scans_by_domain[scan.domain].append(scan)
    for domain in domains:
        domain_scans = scans_by_domain[domain]
        lines += [f"### {domain}", ""]
        extension_counts = Counter(item.extension for item in domain_scans)
        lines.append(
            "附件类型："
            + "、".join(f"{key or '无扩展名'} {value}" for key, value in extension_counts.most_common())
            + "。"
        )
        clusters = defaultdict(list)
        for scan in domain_scans:
            clusters[signature(scan)].append(scan)
        label_presence = Counter(label for scan in domain_scans for label in scan.labels)
        threshold = max(2, math.ceil(len(domain_scans) * 0.15))
        recurring = [] if domain == "163.com" else [
            (label, count)
            for label, count in label_presence.most_common()
            if count >= threshold
        ][:30]
        if recurring:
            lines.append(
                "重复出现的固定标签："
                + "、".join(f"`{label}`（{count}）" for label, count in recurring)
                + "。"
            )
        lines += ["", "| 模板 | 文件数 | 工作表 | 已识别标准字段 | 自动解析判断 |", "|---|---:|---|---|---|"]
        for index, (_, members) in enumerate(
            sorted(clusters.items(), key=lambda item: len(item[1]), reverse=True), 1
        ):
            field_labels = defaultdict(set)
            for member in members:
                for key, labels in member.fields.items():
                    field_labels[key].update(labels)
            fields = "、".join(
                f"{key}（{'/'.join(sorted(labels))}）" for key, labels in sorted(field_labels.items())
            )
            required = set(field_labels)
            if {"unit_nav", "valuation_date"} <= required and required & {
                "product_code",
                "product_name",
            }:
                verdict = "表头字段齐全；产品建档后可按模板自动导入"
            elif field_labels:
                verdict = "部分字段可识别；需补充该托管别名或定位规则"
            else:
                verdict = "当前别名未命中；需按此模板专门适配"
            sheet_names = " / ".join(members[0].sheets) or "—"
            lines.append(
                f"| {domain}-T{index} | {len(members)} | {sheet_names} | {fields or '—'} | {verdict} |"
            )
        errors = Counter(scan.error for scan in domain_scans if scan.error)
        if errors:
            lines += ["", "无法按 Excel 表头读取的附件：" + "、".join(f"{key}（{value}）" for key, value in errors.most_common()) + "。"]
        lines.append("")

    reason_counts = Counter()
    for result in job_results:
        for error in result.get("errors", []):
            reason_counts[error.get("reason", "未知错误")] += 1
    lines += [
        "## 四、当前解析结果归纳",
        "",
        "| 结果或阻塞原因 | 次数 |",
        "|---|---:|",
    ]
    for reason, count in reason_counts.most_common(20):
        lines.append(f"| {reason.replace('|', '｜')} | {count} |")
    lines += [
        "",
        "## 五、建议的模板适配顺序",
        "",
        "1. 先按附件量从高到低适配 `citics.com`、`htsc.com`、`ebscn.com`、`cmschina.com.cn`、`swhysc.com`。",
        "2. 每个托管模板只维护一份字段映射：邮件主题/附件名负责初步识别产品代码与份额，工作簿固定单元格或表头负责净值、日期、规模、份额、持仓和现金字段。",
        "3. 新模板首次出现时进入人工确认；同一模板后续邮件复用解析器，不逐封配置。",
        "4. 产品代码优先匹配；只有产品名称时必须精确匹配。全新产品继续进入“待确认产品”，由运营核对后建档并重新解析。",
        "5. 模板版本以发件域、主题模式、附件扩展名、工作表结构和固定字段位置共同识别；结构变化时生成新的模板版本，不覆盖旧版本。",
        "",
        "## 六、边界说明",
        "",
        "- 本报告只归纳邮件头和附件版式，没有改动原邮件、附件或业务数据。",
        "- 相同托管方可能同时存在净值通知、估值表、对账单等多类模板，应按上表模板编号分别适配。",
        "- PDF、压缩包中非 Excel 文件及加密/损坏工作簿不能仅凭版式自动取数，需要人工确认或独立解析器。",
    ]
    output.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--database", default="sqlite:///runtime/development.db")
    parser.add_argument("--storage", type=Path, default=Path("runtime/archive"))
    parser.add_argument(
        "--output", type=Path, default=Path("docs/邮箱托管模板与可解析表头清单.md")
    )
    args = parser.parse_args()
    engine = create_engine(args.database)
    with engine.connect() as connection:
        originals_raw = connection.execute(
            text(
                "SELECT metadata_json FROM documents "
                "WHERE source='email' AND parent_id IS NULL"
            )
        ).scalars()
        originals = []
        for value in originals_raw:
            item = json.loads(value or "{}")
            originals.append(
                {"from": item.get("from", ""), "subject": item.get("subject", "")}
            )
        rows = connection.execute(
            text(
                "SELECT d.filename,d.storage_key,p.metadata_json "
                "FROM documents d JOIN documents p ON p.id=d.parent_id "
                "WHERE d.source='email' AND d.parent_id IS NOT NULL"
            )
        ).fetchall()
        job_rows = connection.execute(
            text("SELECT status,result FROM parse_jobs")
        ).fetchall()
    engine.dispose()
    attachments = []
    scans = []
    for filename, storage_key, parent_metadata in rows:
        metadata = json.loads(parent_metadata or "{}")
        domain = domain_of(metadata.get("from", ""))
        attachments.append({"filename": filename, "domain": domain})
        content = (args.storage / storage_key).read_bytes()
        found = False
        for nested_name, nested_content in iter_attachment_files(filename, content):
            found = True
            scans.append(scan_book(domain, nested_name, nested_content))
        if not found:
            scans.append(
                Scan(
                    domain=domain,
                    filename=filename,
                    extension=Path(filename).suffix.lower(),
                    error="压缩包中无可读取 Excel" if filename.lower().endswith(".zip") else "非 Excel 附件",
                )
            )
    statuses = Counter(status for status, _ in job_rows)
    job_results = [json.loads(result or "{}") for _, result in job_rows]
    args.output.parent.mkdir(parents=True, exist_ok=True)
    render_report(originals, attachments, scans, statuses, job_results, args.output)
    print(f"Wrote {args.output}: {len(originals)} mails, {len(scans)} scanned files")


if __name__ == "__main__":
    main()
