"""Conservative deterministic import. Unknown layouts are review tasks, never fabricated data."""

import csv
import io
import re
import zipfile
from datetime import date, datetime
from decimal import Decimal, InvalidOperation, localcontext

from openpyxl import load_workbook

ALIASES = {
    "product_code": ["产品代码", "基金代码", "备案编码", "product_code"],
    "product_name": ["产品名称", "基金名称", "product_name"],
    "share_class": ["份额类别", "份额名称", "share_class"],
    "valuation_date": ["估值日期", "净值日期", "业务日期", "valuation_date"],
    "unit_nav": ["单位净值", "基金份额净值", "单位基金净值", "unit_nav"],
    "accumulated_nav": ["累计净值", "累计单位净值", "accumulated_nav"],
    "net_assets": ["资产净值", "基金资产净值", "net_assets"],
    "total_shares": ["份额总数", "基金总份额", "total_shares"],
    "currency": ["币种", "currency"],
    "position_ratio": ["仓位", "仓位比例", "position_ratio"],
    "cash": ["现金", "现金余额", "cash"],
    "return_rate": ["收益率", "区间收益率", "return_rate"],
    "drawdown": ["回撤", "最大回撤", "drawdown"],
}
ALIAS_MAP = {
    re.sub(r"\s+", "", alias).lower(): key
    for key, values in ALIASES.items()
    for alias in values
}


def numeric(value, field):
    if value is None or str(value).strip() in {"", "—", "-", "--"}:
        return None
    if isinstance(value, bool):
        raise ValueError(f"{field} 不是数值")
    text = str(value).strip().replace(",", "").replace("，", "")
    if text.endswith("%"):
        if field not in {"position_ratio", "return_rate", "drawdown"}:
            raise ValueError(f"{field} 不应使用百分比")
        text = str(Decimal(text[:-1]) / 100)
    try:
        result = Decimal(text)
    except InvalidOperation as exc:
        raise ValueError(f"{field} 不是可识别的十进制数") from exc
    integer_limit = (
        Decimal("1e14") if field in {"unit_nav", "accumulated_nav"} else Decimal("1e19")
    )
    precision = (
        10
        if field in {"unit_nav", "accumulated_nav"}
        else 4
        if field in {"net_assets", "total_shares"}
        else 12
    )
    if not result.is_finite() or abs(result) >= integer_limit:
        raise ValueError(f"{field} 超出允许范围")
    with localcontext() as context:
        context.prec = 50
        if result != result.quantize(Decimal(1).scaleb(-precision)):
            raise ValueError(f"{field} 小数位超过 {precision} 位，禁止静默舍入")
        return format(result.normalize(), "f")


def business_date(value):
    if isinstance(value, (date, datetime)):
        return value.strftime("%Y-%m-%d")
    text = str(value).strip().replace("/", "-").replace(".", "-")
    text = re.sub(r"年", "-", text)
    text = re.sub(r"月", "-", text)
    text = re.sub(r"日$", "", text)
    if re.fullmatch(r"\d{8}", text):
        text = f"{text[:4]}-{text[4:6]}-{text[6:]}"
    return date.fromisoformat(text).isoformat()


def normalize_record(row):
    result = {}
    for key, value in row.items():
        if value is None or str(value).strip() == "":
            continue
        if key in {"product_code", "product_name", "share_class", "currency"}:
            result[key] = str(value).strip()
        elif key == "valuation_date":
            result[key] = business_date(value)
        else:
            result[key] = numeric(value, key)
    if not result.get("unit_nav") or not result.get("valuation_date"):
        raise ValueError("缺少单位净值或估值日期")
    if not result.get("product_code") and not result.get("product_name"):
        raise ValueError("缺少产品名称或产品代码")
    return result


def extract_tables(filename, content, max_rows=10000):
    suffix = filename.lower().rsplit(".", 1)[-1]
    if suffix in {"xlsx", "xlsm"}:
        if suffix == "xlsm":
            raise ValueError("暂不接收含宏工作簿")
        with zipfile.ZipFile(io.BytesIO(content)) as z:
            if (
                len(z.infolist()) > 2000
                or sum(f.file_size for f in z.infolist()) > 80 * 1024 * 1024
            ):
                raise ValueError("工作簿解压大小或文件数量超限")
        book = load_workbook(
            io.BytesIO(content), read_only=True, data_only=True, keep_links=False
        )
        try:
            if len(book.sheetnames) > 32:
                raise ValueError("工作表数量超限")
            total_rows = 0
            for sheet in book:
                if (sheet.max_row or 0) > max_rows or (sheet.max_column or 0) > 200:
                    raise ValueError("工作表超出行列限制")
                rows = []
                for i, row in enumerate(sheet.iter_rows(values_only=True)):
                    total_rows += 1
                    if total_rows > max_rows:
                        raise ValueError("工作簿总行数超限")
                    if i >= max_rows or len(row) > 200:
                        raise ValueError("工作表超出行列限制")
                    rows.append(row)
                yield sheet.title, rows
        finally:
            book.close()
    elif suffix == "xls":
        import xlrd

        book = xlrd.open_workbook(file_contents=content, on_demand=True)
        try:
            for sheet in book.sheets():
                if sheet.nrows > max_rows or sheet.ncols > 200:
                    raise ValueError("工作表超出行列限制")
                rows = []
                for index in range(sheet.nrows):
                    rows.append(
                        [
                            xlrd.xldate_as_datetime(c.value, book.datemode)
                            if c.ctype == xlrd.XL_CELL_DATE
                            else c.value
                            for c in sheet.row(index)
                        ]
                    )
                yield sheet.name, rows
        finally:
            book.release_resources()
    elif suffix == "csv":
        try:
            text = content.decode("utf-8-sig")
        except UnicodeDecodeError:
            text = content.decode("gb18030")
        rows = []
        for i, row in enumerate(csv.reader(io.StringIO(text))):
            if i >= max_rows or len(row) > 200:
                raise ValueError("CSV 超出行列限制")
            rows.append(row)
        yield "CSV", rows
    elif suffix == "pdf":
        # Archive PDFs, but never infer financial fields from unverified PDF layouts.
        raise ValueError("PDF 已归档，需核验托管模板后适配；可人工补录净值")
    else:
        raise ValueError("该格式已归档，尚无经过验证的解析器")


def _generic_parse(tables):
    records, errors = [], []
    for sheet, rows in tables:
        header = None
        indices = {}
        for row_number, cells in enumerate(rows, 1):
            mapping = {
                i: ALIAS_MAP.get(re.sub(r"\s+", "", str(v)).lower())
                for i, v in enumerate(cells)
                if v is not None
            }
            mapping = {i: v for i, v in mapping.items() if v}
            if "unit_nav" in mapping.values() and "valuation_date" in mapping.values():
                if len(set(mapping.values())) != len(mapping):
                    raise ValueError("同一表头含重复指标，不能自动判断正确列")
                header, indices = row_number, mapping
                continue
            if header and any(v is not None and str(v).strip() for v in cells):
                row = {
                    field: cells[i] for i, field in indices.items() if i < len(cells)
                }
                if not row.get("unit_nav") and not row.get("valuation_date"):
                    continue
                try:
                    records.append(
                        {**normalize_record(row), "row_key": f"{sheet}:{row_number}"}
                    )
                except (ValueError, InvalidOperation) as exc:
                    errors.append({"row": f"{sheet}:{row_number}", "reason": str(exc)})
        if not header:
            # Explicit key/value sheets; no positional guesses or fuzzy field matches.
            values = {}
            for row in rows:
                if len(row) >= 2:
                    key = ALIAS_MAP.get(re.sub(r"\s+", "", str(row[0])).lower())
                    if key:
                        if key in values:
                            errors.append(
                                {
                                    "row": sheet,
                                    "reason": "重复字段，可能包含多个份额，需模板适配",
                                }
                            )
                            values = {}
                            break
                        values[key] = row[1]
            if values.get("unit_nav"):
                try:
                    records.append(
                        {**normalize_record(values), "row_key": f"{sheet}:kv"}
                    )
                except (ValueError, InvalidOperation) as exc:
                    errors.append({"row": sheet, "reason": str(exc)})
    if not records and not errors:
        errors.append(
            {
                "row": "file",
                "reason": "未发现可识别净值表头；原件已保留，需适配模板或人工补录",
            }
        )
    return records, errors


def _parse_workbook(filename, content, max_rows, context):
    from .custodian_parsers import is_known_custodian, select_adapter

    tables = list(extract_tables(filename, content, max_rows))
    adapter = select_adapter(tables, context)
    if not adapter:
        if is_known_custodian(context):
            return (
                [],
                [
                    {
                        "row": "file",
                        "reason": "已识别托管方，但附件结构与已验证模板不一致，需人工确认并新增模板版本",
                    }
                ],
                "custodian:unmatched:1",
            )
        records, errors = _generic_parse(tables)
        return records, errors, "explicit-header-v2"
    raw = adapter.parse(filename, tables, context)
    records, errors = [], list(raw.errors)
    for item in raw.records:
        row_key = item.pop("row_key", "file")
        try:
            records.append({**normalize_record(item), "row_key": row_key})
        except (ValueError, InvalidOperation) as exc:
            errors.append({"row": row_key, "reason": str(exc)})
    if not records and not errors:
        errors.append({"row": "file", "reason": "托管模板已识别，但未发现有效净值记录"})
    return records, errors, adapter.parser_version


def parse(filename, content, max_rows=10000, context=None):
    """Parse a workbook or a safe ZIP and return canonical NAV records.

    ``context`` may contain sender_domain, sender, subject and source. It improves
    adapter selection but is not required: manual uploads use workbook signatures.
    """
    context = dict(context or {})
    if filename.lower().endswith(".zip"):
        records, errors, versions = [], [], set()
        try:
            archive = zipfile.ZipFile(io.BytesIO(content))
        except zipfile.BadZipFile as exc:
            raise ValueError("压缩包损坏，无法读取") from exc
        with archive:
            members = [member for member in archive.infolist() if not member.is_dir()]
            if len(members) > 100 or sum(member.file_size for member in members) > 100 * 1024 * 1024:
                raise ValueError("压缩包解压大小或文件数量超限")
            for member in members:
                member_name = member.filename.replace("\\", "/").rsplit("/", 1)[-1]
                if not member_name.lower().endswith((".xlsx", ".xls", ".csv")):
                    continue
                member_records, member_errors, version = _parse_workbook(
                    member_name, archive.read(member), max_rows, context
                )
                for record in member_records:
                    record["row_key"] = f"{member_name}:{record['row_key']}"
                for error in member_errors:
                    error["row"] = f"{member_name}:{error.get('row', 'file')}"
                records.extend(member_records)
                errors.extend(member_errors)
                versions.add(version)
        if not records and not errors:
            errors.append({"row": "file", "reason": "压缩包内没有可解析的 Excel/CSV 附件"})
        return {
            "records": records,
            "errors": errors,
            "parser_version": "+".join(sorted(versions)) or "archive-v1",
        }
    records, errors, version = _parse_workbook(filename, content, max_rows, context)
    return {"records": records, "errors": errors, "parser_version": version}
